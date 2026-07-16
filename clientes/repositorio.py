"""Lectura y escritura de la base de datos de clientes.

Es un archivo Excel local (`datos/clientes.xlsx`, con formato — ver
`clientes/generar_plantilla.py`) que Fernando edita a mano (elige el
programa de un desplegable; tarifa y sesiones totales se calculan solas) y
que este módulo lee y actualiza tras cada resumen semanal. No depende de
ningún conector ni credencial: es un archivo del propio ordenador. Al
escribir solo se cambian valores de celda, nunca el formato, así que el
aspecto del Excel no se pierde.

Importante: la tarifa y las sesiones totales son fórmulas de Excel, no
valores fijos. Este módulo lee el archivo con `data_only=True`, que devuelve
el último valor calculado por Excel — por eso Fernando debe **guardar el
archivo (Ctrl+S) después de elegir un programa** para que el sistema pueda
leer esos números.
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from programas.logica import ActualizacionPrograma

RUTA_POR_DEFECTO = Path(__file__).resolve().parent.parent / "datos" / "clientes.xlsx"
HOJA = "Clientes"
HOJA_PROGRAMAS = "Programas"
PRIMERA_FILA_DATOS = 3
ULTIMA_FILA_CON_HUECO = 30


def _tabla_programas(wb) -> dict[str, tuple]:
    """Lee la hoja "Programas" (valores literales, nunca fórmulas) como
    {nombre_programa: (tarifa, sesiones_totales)} — sirve de respaldo
    cuando el valor calculado de la fórmula en "Clientes" no está
    disponible (ver más abajo)."""
    hoja = wb[HOJA_PROGRAMAS]
    tabla: dict[str, tuple] = {}
    fila = 3
    while hoja[f"A{fila}"].value:
        tabla[hoja[f"A{fila}"].value] = (hoja[f"B{fila}"].value, hoja[f"C{fila}"].value)
        fila += 1
    return tabla


def leer_clientes(ruta: Path = RUTA_POR_DEFECTO) -> dict[str, dict]:
    """Devuelve {cliente: {fila, tipo_programa, tarifa, sesiones_totales,
    sesiones_llevadas, pendiente_pago}} tal cual está en el Excel.

    Fernando anota las sesiones "llevadas" (consumidas del bono actual), no
    las que le quedan — así lo pidió el 2026-07-15. `a_programa` hace la
    conversión a "restantes" para la lógica de `programas`.

    Tarifa y sesiones totales son fórmulas en el Excel (VLOOKUP contra la
    hoja "Programas"); si Excel no ha recalculado y guardado el archivo
    (por ejemplo, justo después de que este mismo código haya escrito algo),
    el valor cacheado puede venir vacío. En ese caso se recalcula aquí
    mismo, en Python, contra la misma hoja "Programas" — así no dependemos
    de que Fernando reabra el Excel y pulse Ctrl+S para que el sistema siga
    funcionando (lección del 2026-07-15/2026-07-16).
    """
    wb = load_workbook(ruta, data_only=True)
    hoja = wb[HOJA]
    tabla_programas = _tabla_programas(wb)

    clientes: dict[str, dict] = {}
    fila = PRIMERA_FILA_DATOS
    while hoja[f"A{fila}"].value:
        tipo_programa = hoja[f"B{fila}"].value
        tarifa = hoja[f"C{fila}"].value
        sesiones_totales = hoja[f"D{fila}"].value

        if (tarifa is None or sesiones_totales is None) and tipo_programa in tabla_programas:
            tarifa_respaldo, sesiones_respaldo = tabla_programas[tipo_programa]
            tarifa = tarifa if tarifa is not None else tarifa_respaldo
            sesiones_totales = sesiones_totales if sesiones_totales is not None else sesiones_respaldo

        clientes[hoja[f"A{fila}"].value] = {
            "fila": fila,
            "tipo_programa": tipo_programa,
            "tarifa": tarifa,
            "sesiones_totales": sesiones_totales,
            "sesiones_llevadas": hoja[f"E{fila}"].value,
            "pendiente_pago": hoja[f"F{fila}"].value,
        }
        fila += 1

    return clientes


def a_programa(fila: dict) -> dict | None:
    """Convierte una fila en el formato que espera `programas.procesar`
    (que trabaja en "sesiones restantes", no "llevadas").

    Devuelve None si al cliente le faltan datos por rellenar (tarifa,
    sesiones totales, etc.) — así se puede avisar a Fernando en vez de
    calcular con números inventados.
    """
    try:
        sesiones_totales = int(fila["sesiones_totales"])
        sesiones_llevadas = int(fila["sesiones_llevadas"])
        return {
            "sesiones_restantes": sesiones_totales - sesiones_llevadas,
            "sesiones_totales": sesiones_totales,
            "pendiente_pago": str(fila["pendiente_pago"]).strip().lower() in ("sí", "si"),
        }
    except (TypeError, ValueError):
        return None


def cargar_programas(ruta: Path = RUTA_POR_DEFECTO) -> tuple[dict[str, dict], list[str]]:
    """Lee el Excel y lo deja listo para `programas.procesar.procesar_semana`.

    Devuelve (programas, incompletos): los clientes sin tarifa/sesiones
    rellenas todavía se listan aparte en vez de calcular con datos inventados.
    """
    clientes = leer_clientes(ruta)
    programas: dict[str, dict] = {}
    incompletos: list[str] = []

    for nombre, fila in clientes.items():
        programa = a_programa(fila)
        if programa is None:
            incompletos.append(nombre)
        else:
            programas[nombre] = programa

    return programas, incompletos


def cargar_tarifas(ruta: Path = RUTA_POR_DEFECTO) -> dict[str, float]:
    """Devuelve {cliente: tarifa} para los clientes con tarifa numérica ya
    calculada — usado por `economia.calculo` para la facturación semanal."""
    clientes = leer_clientes(ruta)
    tarifas: dict[str, float] = {}
    for nombre, fila in clientes.items():
        try:
            tarifas[nombre] = float(fila["tarifa"])
        except (TypeError, ValueError):
            continue
    return tarifas


def _asegurar_validaciones(wb) -> None:
    """Repone los desplegables si no están (openpyxl no lee el formato
    "extendido" en el que Excel a veces reescribe las validaciones al
    guardar, y los descarta al reabrir el archivo — ver lección del
    2026-07-15 en el log). Se comprueba y repone en cada escritura para que
    el desplegable nunca desaparezca sin que nadie se dé cuenta."""
    hoja = wb[HOJA]
    validaciones = hoja.data_validations.dataValidation

    tiene_validacion_programa = any("Programas!" in (dv.formula1 or "") for dv in validaciones)
    tiene_validacion_pago = any(dv.formula1 == '"Sí,No"' for dv in validaciones)

    if not tiene_validacion_programa:
        hoja_programas = wb[HOJA_PROGRAMAS]
        ultima_fila_programas = 2
        while hoja_programas[f"A{ultima_fila_programas + 1}"].value:
            ultima_fila_programas += 1
        validacion = DataValidation(
            type="list", formula1=f"=Programas!$A$3:$A${ultima_fila_programas}", allow_blank=True
        )
        hoja.add_data_validation(validacion)
        validacion.add(f"B{PRIMERA_FILA_DATOS}:B{ULTIMA_FILA_CON_HUECO}")

    if not tiene_validacion_pago:
        validacion = DataValidation(type="list", formula1='"Sí,No"', allow_blank=False)
        hoja.add_data_validation(validacion)
        validacion.add(f"F{PRIMERA_FILA_DATOS}:F{ULTIMA_FILA_CON_HUECO}")


def actualizar_cliente(
    nombre: str, sesiones_llevadas: int, pendiente_pago: bool, ruta: Path = RUTA_POR_DEFECTO
) -> None:
    """Edición manual de un cliente concreto (usada por la web app):
    escribe sesiones llevadas y pendiente de pago directamente, sin pasar
    por la lógica de renovación de `programas.procesar` — es una
    corrección puntual, no un cierre semanal."""
    clientes = leer_clientes(ruta)
    if nombre not in clientes:
        raise ValueError(f"No existe el cliente '{nombre}'")

    fila = clientes[nombre]["fila"]
    wb = load_workbook(ruta)
    hoja = wb[HOJA]
    hoja[f"E{fila}"] = sesiones_llevadas
    hoja[f"F{fila}"] = "Sí" if pendiente_pago else "No"

    _asegurar_validaciones(wb)
    wb.save(ruta)


def aplicar_actualizaciones(
    resultados: dict[str, ActualizacionPrograma], ruta: Path = RUTA_POR_DEFECTO
) -> None:
    """Escribe en el Excel las sesiones llevadas y el pendiente de pago ya
    calculados (convirtiendo de "restantes" a "llevadas"). Solo se llama
    después de que Fernando confirme el resumen. Solo se tocan valores de
    celda: el formato del Excel no cambia."""
    clientes = leer_clientes(ruta)
    wb = load_workbook(ruta)
    hoja = wb[HOJA]

    for nombre, actualizacion in resultados.items():
        fila = clientes[nombre]["fila"]
        sesiones_totales = int(clientes[nombre]["sesiones_totales"])
        hoja[f"E{fila}"] = sesiones_totales - actualizacion.sesiones_restantes
        hoja[f"F{fila}"] = "Sí" if actualizacion.pendiente_pago else "No"

    _asegurar_validaciones(wb)
    wb.save(ruta)
