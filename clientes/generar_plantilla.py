"""Genera datos/clientes.xlsx: un Excel con formato para que Fernando rellene
los datos de sus clientes a mano. Se ejecuta una sola vez (o cuando se quiera
regenerar la plantilla desde cero); las ediciones normales de Fernando se
hacen abriendo el archivo directamente en Excel.

Incluye una hoja "Programas" con las tarifas reales de Antifrágil (ver
docs/TARIFAS.md): al elegir un programa en la hoja "Clientes", la tarifa y
las sesiones totales se rellenan solas mediante una fórmula de búsqueda. Si
Fernando cambia un precio en el futuro, solo tiene que editar la hoja
"Programas" — no hace falta tocar código.

Uso: python -m clientes.generar_plantilla
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RUTA_SALIDA = Path(__file__).resolve().parent.parent / "datos" / "clientes.xlsx"

CLIENTES_INICIALES = [
    "Nikki", "Felipe y Javi", "Rocío", "Samanta", "Neha", "Paquito", "Ana", "Sunil",
]

# (nombre del programa, tarifa €, sesiones totales) — ver docs/TARIFAS.md
PROGRAMAS = [
    ("Nuevo 45€ x4", 45, 4),
    ("Antiguo 40€ x4", 40, 4),
    ("Nuevo 40€ x8", 40, 8),
    ("Antiguo 35€ x8", 35, 8),
    ("Antiguo 35€ x16", 35, 16),
    ("Nuevo 37,50€ x16", 37.5, 16),
    ("Pareja 60€ x12", 60, 12),
]

COLUMNAS = [
    ("Cliente", 26),
    ("Tipo de programa", 22),
    ("Tarifa (€)", 14),
    ("Sesiones totales", 16),
    ("Sesiones restantes", 18),
    ("Pendiente de pago", 18),
]

FILAS_EXTRA_PARA_CLIENTES_NUEVOS = 20

AZUL_ANTIFRAGIL = "1F3B4D"
GRIS_CLARO = "F2F2F2"
VERDE_SUAVE = "C6EFCE"
VERDE_TEXTO = "006100"
ROJO_SUAVE = "FFC7CE"
ROJO_TEXTO = "9C0006"

BORDE_FINO = Border(*(Side(style="thin", color="D9D9D9") for _ in range(4)))


def _estilar_cabecera(hoja, fila_cabecera, columnas):
    for col_idx, (nombre, ancho) in enumerate(columnas, start=1):
        letra = get_column_letter(col_idx)
        celda = hoja[f"{letra}{fila_cabecera}"]
        celda.value = nombre
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=AZUL_ANTIFRAGIL)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = BORDE_FINO
        hoja.column_dimensions[letra].width = ancho
    hoja.row_dimensions[fila_cabecera].height = 24


def _construir_hoja_programas(wb) -> None:
    hoja = wb.create_sheet("Programas")
    hoja["A1"] = "Tabla de programas y tarifas (edita aquí si cambian los precios)"
    hoja.merge_cells("A1:C1")
    hoja["A1"].font = Font(size=11, bold=True, color="FFFFFF")
    hoja["A1"].fill = PatternFill("solid", fgColor=AZUL_ANTIFRAGIL)
    hoja["A1"].alignment = Alignment(horizontal="center", vertical="center")

    _estilar_cabecera(hoja, 2, [("Programa", 22), ("Tarifa (€)", 14), ("Sesiones", 12)])

    for i, (nombre, tarifa, sesiones) in enumerate(PROGRAMAS):
        fila = 3 + i
        hoja[f"A{fila}"] = nombre
        hoja[f"B{fila}"] = tarifa
        hoja[f"B{fila}"].number_format = "#,##0.00 €"
        hoja[f"C{fila}"] = sesiones
        for letra in ("A", "B", "C"):
            hoja[f"{letra}{fila}"].border = BORDE_FINO
            hoja[f"{letra}{fila}"].alignment = Alignment(horizontal="center", vertical="center")
        hoja[f"A{fila}"].alignment = Alignment(horizontal="left", vertical="center")


def construir_workbook() -> Workbook:
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Clientes"

    _construir_hoja_programas(wb)

    # Título
    hoja.merge_cells("A1:F1")
    titulo = hoja["A1"]
    titulo.value = "Antifrágil — Clientes de Entrenamiento Personal"
    titulo.font = Font(size=14, bold=True, color="FFFFFF")
    titulo.fill = PatternFill("solid", fgColor=AZUL_ANTIFRAGIL)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    hoja.row_dimensions[1].height = 28

    fila_cabecera = 2
    _estilar_cabecera(hoja, fila_cabecera, COLUMNAS)

    primera_fila_datos = fila_cabecera + 1
    ultima_fila_clientes = primera_fila_datos + len(CLIENTES_INICIALES) - 1
    ultima_fila_con_hueco = ultima_fila_clientes + FILAS_EXTRA_PARA_CLIENTES_NUEVOS

    # Filas de datos (clientes ya detectados, más una zona en blanco para clientes nuevos)
    for fila in range(primera_fila_datos, ultima_fila_con_hueco + 1):
        i = fila - primera_fila_datos
        if i < len(CLIENTES_INICIALES):
            hoja[f"A{fila}"] = CLIENTES_INICIALES[i]

        for col_idx in range(1, len(COLUMNAS) + 1):
            letra = get_column_letter(col_idx)
            celda = hoja[f"{letra}{fila}"]
            celda.border = BORDE_FINO
            if i % 2 == 1:
                celda.fill = PatternFill("solid", fgColor=GRIS_CLARO)
            celda.alignment = Alignment(horizontal="center", vertical="center")
        hoja[f"A{fila}"].alignment = Alignment(horizontal="left", vertical="center")

        # Tarifa y sesiones totales se calculan solas a partir del programa elegido
        hoja[f"C{fila}"] = f'=IFERROR(VLOOKUP($B{fila},Programas!$A:$C,2,FALSE),"")'
        hoja[f"C{fila}"].number_format = "#,##0.00 €"
        hoja[f"D{fila}"] = f'=IFERROR(VLOOKUP($B{fila},Programas!$A:$C,3,FALSE),"")'
        hoja[f"D{fila}"].number_format = "0"
        hoja[f"E{fila}"].number_format = "0"
        if i < len(CLIENTES_INICIALES):
            hoja[f"F{fila}"] = "No"

    # Desplegable con los programas reales para "Tipo de programa"
    validacion_programa = DataValidation(
        type="list", formula1=f"=Programas!$A$3:$A${2 + len(PROGRAMAS)}", allow_blank=True
    )
    hoja.add_data_validation(validacion_programa)
    validacion_programa.add(f"B{primera_fila_datos}:B{ultima_fila_con_hueco}")

    # Desplegable Sí/No para "Pendiente de pago"
    validacion_pago = DataValidation(type="list", formula1='"Sí,No"', allow_blank=False)
    hoja.add_data_validation(validacion_pago)
    validacion_pago.add(f"F{primera_fila_datos}:F{ultima_fila_con_hueco}")

    # Resaltado de pendientes de pago
    hoja.conditional_formatting.add(
        f"F{primera_fila_datos}:F{ultima_fila_con_hueco}",
        CellIsRule(operator="equal", formula=['"Sí"'], fill=PatternFill("solid", fgColor=ROJO_SUAVE),
                   font=Font(color=ROJO_TEXTO)),
    )
    hoja.conditional_formatting.add(
        f"F{primera_fila_datos}:F{ultima_fila_con_hueco}",
        CellIsRule(operator="equal", formula=['"No"'], fill=PatternFill("solid", fgColor=VERDE_SUAVE),
                   font=Font(color=VERDE_TEXTO)),
    )

    hoja.freeze_panes = f"A{primera_fila_datos}"
    hoja.auto_filter.ref = f"A{fila_cabecera}:F{ultima_fila_clientes}"

    wb.active = 0
    return wb


def main() -> None:
    wb = construir_workbook()
    wb.save(RUTA_SALIDA)
    print(f"Plantilla creada en {RUTA_SALIDA}")


if __name__ == "__main__":
    main()
