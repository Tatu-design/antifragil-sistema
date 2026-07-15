"""Registro histórico de facturación (`datos/facturacion.xlsx`), consultable
por semana o por mes en cualquier momento.

Tres hojas:
- "Semanas": una fila por semana cerrada (facturación, horas, precio medio).
- "Desglose": una fila por (semana, tarifa) — el detalle que antes llevaba
  Fernando a mano, ahora en formato de tabla en vez de celdas fusionadas.
- "Meses": una fila por mes, agregando las semanas cuyo lunes cae en ese mes.

CrossFit Kids se registra sin facturación hasta que Fernando indica el
importe mensual (`registrar_facturacion_kids`), momento en el que se
reparte hacia atrás sobre las semanas de ese mes (importe ÷ sesiones del
mes = precio por sesión; cada semana se multiplica por sus sesiones).
"""

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from economia.calculo import resumir

RUTA_POR_DEFECTO = Path(__file__).resolve().parent.parent / "datos" / "facturacion.xlsx"

HOJA_SEMANAS = "Semanas"
HOJA_DESGLOSE = "Desglose"
HOJA_MESES = "Meses"

CABECERA_SEMANAS = [
    "Fecha inicio", "Fecha fin", "Año", "Mes",
    "Facturación PT+Lidomare", "Horas PT+Lidomare", "Precio medio hora",
    "Sesiones CrossFit Kids", "Facturación CrossFit Kids",
    "Facturación Total", "Horas Totales", "Precio medio hora Total",
]
CABECERA_DESGLOSE = ["Fecha inicio semana", "Tarifa (€)", "Sesiones", "Facturación"]
CABECERA_MESES = [
    "Año", "Mes", "Facturación Total", "Horas Totales", "Precio medio hora",
    "Sesiones CrossFit Kids", "Facturación CrossFit Kids",
]

AZUL_ANTIFRAGIL = "1F3B4D"


def _estilar_cabecera(hoja, cabecera: list[str]) -> None:
    for col, nombre in enumerate(cabecera, start=1):
        celda = hoja.cell(row=1, column=col, value=nombre)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=AZUL_ANTIFRAGIL)


def _crear_libro() -> Workbook:
    wb = Workbook()
    hoja = wb.active
    hoja.title = HOJA_SEMANAS
    _estilar_cabecera(hoja, CABECERA_SEMANAS)

    hoja_desglose = wb.create_sheet(HOJA_DESGLOSE)
    _estilar_cabecera(hoja_desglose, CABECERA_DESGLOSE)

    hoja_meses = wb.create_sheet(HOJA_MESES)
    _estilar_cabecera(hoja_meses, CABECERA_MESES)

    return wb


def _abrir_o_crear(ruta: Path) -> Workbook:
    if ruta.exists():
        return load_workbook(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    return _crear_libro()


def _proxima_fila_libre(hoja) -> int:
    fila = 2
    while hoja.cell(row=fila, column=1).value is not None:
        fila += 1
    return fila


def _buscar_fila(hoja, valor_clave, columna: int = 1) -> int | None:
    fila = 2
    while hoja.cell(row=fila, column=1).value is not None:
        if hoja.cell(row=fila, column=columna).value == valor_clave:
            return fila
        fila += 1
    return None


def _buscar_fila_mes(hoja, anio: int, mes: int) -> int | None:
    fila = 2
    while hoja.cell(row=fila, column=1).value is not None:
        if hoja.cell(row=fila, column=1).value == anio and hoja.cell(row=fila, column=2).value == mes:
            return fila
        fila += 1
    return None


def _reemplazar_desglose_semana(hoja, clave_fecha: str, desglose: dict[float, dict]) -> None:
    conservadas = []
    fila = 2
    while hoja.cell(row=fila, column=1).value is not None:
        if hoja.cell(row=fila, column=1).value != clave_fecha:
            conservadas.append([hoja.cell(row=fila, column=c).value for c in range(1, 5)])
        fila += 1

    for fila in range(2, hoja.max_row + 1):
        for col in range(1, 5):
            hoja.cell(row=fila, column=col, value=None)

    fila_actual = 2
    for valores in conservadas:
        for col, valor in enumerate(valores, start=1):
            hoja.cell(row=fila_actual, column=col, value=valor)
        fila_actual += 1

    for tarifa, datos in sorted(desglose.items()):
        hoja.cell(row=fila_actual, column=1, value=clave_fecha)
        hoja.cell(row=fila_actual, column=2, value=tarifa)
        hoja.cell(row=fila_actual, column=3, value=datos["sesiones"])
        hoja.cell(row=fila_actual, column=4, value=datos["facturacion"])
        fila_actual += 1


def _actualizar_mes(wb: Workbook, anio: int, mes: int) -> None:
    hoja_semanas = wb[HOJA_SEMANAS]
    hoja_meses = wb[HOJA_MESES]

    facturacion_total = 0.0
    horas_totales = 0
    sesiones_kids = 0
    facturacion_kids = 0.0
    kids_conocido = False

    fila = 2
    while hoja_semanas.cell(row=fila, column=1).value is not None:
        if hoja_semanas.cell(row=fila, column=3).value == anio and hoja_semanas.cell(row=fila, column=4).value == mes:
            facturacion_total += hoja_semanas.cell(row=fila, column=10).value or 0
            horas_totales += hoja_semanas.cell(row=fila, column=11).value or 0
            sesiones_kids += hoja_semanas.cell(row=fila, column=8).value or 0
            fk = hoja_semanas.cell(row=fila, column=9).value
            if isinstance(fk, (int, float)):
                facturacion_kids += fk
                kids_conocido = True
        fila += 1

    precio_medio = facturacion_total / horas_totales if horas_totales else 0.0
    fila_mes = _buscar_fila_mes(hoja_meses, anio, mes) or _proxima_fila_libre(hoja_meses)
    valores = [
        anio, mes, facturacion_total, horas_totales, precio_medio,
        sesiones_kids, facturacion_kids if kids_conocido else None,
    ]
    for col, valor in enumerate(valores, start=1):
        hoja_meses.cell(row=fila_mes, column=col, value=valor)


def registrar_semana(
    fecha_inicio: date,
    fecha_fin: date,
    desglose: dict[float, dict],
    sesiones_kids: int,
    ruta: Path = RUTA_POR_DEFECTO,
) -> None:
    """Guarda (o actualiza, si ya existía) el resultado económico de una semana."""
    wb = _abrir_o_crear(ruta)
    hoja_semanas = wb[HOJA_SEMANAS]
    hoja_desglose = wb[HOJA_DESGLOSE]

    resumen = resumir(desglose)
    anio, mes = fecha_inicio.year, fecha_inicio.month
    clave = fecha_inicio.isoformat()

    fila = _buscar_fila(hoja_semanas, clave) or _proxima_fila_libre(hoja_semanas)

    facturacion_kids_previa = hoja_semanas.cell(row=fila, column=9).value
    facturacion_kids = facturacion_kids_previa if isinstance(facturacion_kids_previa, (int, float)) else None

    facturacion_total = resumen["facturacion_total"] + (facturacion_kids or 0)
    horas_totales = resumen["horas_totales"] + sesiones_kids
    precio_medio_total = facturacion_total / horas_totales if horas_totales else 0.0

    valores = [
        clave, fecha_fin.isoformat(), anio, mes,
        resumen["facturacion_total"], resumen["horas_totales"], resumen["precio_medio_hora"],
        sesiones_kids, facturacion_kids,
        facturacion_total, horas_totales, precio_medio_total,
    ]
    for col, valor in enumerate(valores, start=1):
        hoja_semanas.cell(row=fila, column=col, value=valor)

    _reemplazar_desglose_semana(hoja_desglose, clave, desglose)
    _actualizar_mes(wb, anio, mes)
    wb.save(ruta)


def registrar_facturacion_kids(anio: int, mes: int, facturacion_total_kids: float, ruta: Path = RUTA_POR_DEFECTO) -> float:
    """Reparte la facturación mensual de CrossFit Kids entre las semanas de
    ese mes, proporcionalmente a las sesiones de cada semana. Devuelve el
    precio por sesión calculado."""
    wb = _abrir_o_crear(ruta)
    hoja_semanas = wb[HOJA_SEMANAS]

    filas_mes = []
    sesiones_kids_mes = 0
    fila = 2
    while hoja_semanas.cell(row=fila, column=1).value is not None:
        if hoja_semanas.cell(row=fila, column=3).value == anio and hoja_semanas.cell(row=fila, column=4).value == mes:
            filas_mes.append(fila)
            sesiones_kids_mes += hoja_semanas.cell(row=fila, column=8).value or 0
        fila += 1

    if sesiones_kids_mes == 0:
        raise ValueError(f"No hay sesiones de CrossFit Kids registradas para {mes}/{anio}.")

    precio_sesion = facturacion_total_kids / sesiones_kids_mes

    for fila in filas_mes:
        sesiones_semana = hoja_semanas.cell(row=fila, column=8).value or 0
        facturacion_kids_semana = sesiones_semana * precio_sesion
        facturacion_pt = hoja_semanas.cell(row=fila, column=5).value or 0
        horas_pt = hoja_semanas.cell(row=fila, column=6).value or 0

        nueva_facturacion_total = facturacion_pt + facturacion_kids_semana
        nuevas_horas_totales = horas_pt + sesiones_semana

        hoja_semanas.cell(row=fila, column=9, value=facturacion_kids_semana)
        hoja_semanas.cell(row=fila, column=10, value=nueva_facturacion_total)
        hoja_semanas.cell(row=fila, column=11, value=nuevas_horas_totales)
        hoja_semanas.cell(
            row=fila, column=12,
            value=nueva_facturacion_total / nuevas_horas_totales if nuevas_horas_totales else 0.0,
        )

    _actualizar_mes(wb, anio, mes)
    wb.save(ruta)
    return precio_sesion


def _fila_a_dict(hoja, fila: int, cabecera: list[str]) -> dict:
    return {nombre: hoja.cell(row=fila, column=col).value for col, nombre in enumerate(cabecera, start=1)}


def obtener_semana(fecha_inicio_iso: str, ruta: Path = RUTA_POR_DEFECTO) -> dict | None:
    if not ruta.exists():
        return None
    wb = load_workbook(ruta, data_only=True)
    hoja = wb[HOJA_SEMANAS]
    fila = _buscar_fila(hoja, fecha_inicio_iso)
    return _fila_a_dict(hoja, fila, CABECERA_SEMANAS) if fila else None


def obtener_mes(anio: int, mes: int, ruta: Path = RUTA_POR_DEFECTO) -> dict | None:
    if not ruta.exists():
        return None
    wb = load_workbook(ruta, data_only=True)
    hoja = wb[HOJA_MESES]
    fila = _buscar_fila_mes(hoja, anio, mes)
    return _fila_a_dict(hoja, fila, CABECERA_MESES) if fila else None
